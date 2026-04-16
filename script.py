import streamlit as st
import pandas as pd
import io
import datetime
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ========================
# 計算函式
# ========================
def calculate_bonus(deal_counts, extra_classes, loyalty_counts, upgrade_counts, is_full_time, brand_count):
    # 1. 體驗成交獎金
    d_total = (deal_counts.get("當天", 0) * 80 + 
               deal_counts.get("48小時", 0) * 60 + 
               deal_counts.get("7天內", 0) * 50)
    # 2. 補位獎金
    c_total = extra_classes * 30
    # 3. 回流獎勵獎金 (STP-T)
    l_total = (loyalty_counts.get("10堂", 0) * 100 + 
               loyalty_counts.get("20堂", 0) * 200 + 
               loyalty_counts.get("30堂", 0) * 300 + 
               loyalty_counts.get("40堂", 0) * 500)
    # 4. 結構升級獎金
    upgrade_prices = {"1對2變1對3": 100, "團課變期班": 150, "包班成立": 300}
    u_total = sum(upgrade_prices.get(name, 0) * count for name, count in upgrade_counts.items())

    # 5. 品牌知名度提升獎金與備註
    base_val = 5 if is_full_time else 2
    b_note = ""
    if brand_count == 0:
        b_total = -200
        b_note = "推廣人數為 0"
    elif brand_count < base_val:
        b_total = -100
        b_note = f"未達門檻 ({base_val}位)"
    elif brand_count == base_val:
        b_total = 0
        b_note = "符合基本門檻"
    else:
        extra_units = (brand_count - base_val) // 5
        b_total = extra_units * 200
        b_note = f"加發 {extra_units} 組獎金"

    # 6. 月轉換高手筆數累計
    total_deals = (sum(deal_counts.values()) + sum(upgrade_counts.values()) + 
                   sum(loyalty_counts.values()) + extra_classes)
    monthly_bonus = 5000 if total_deals >= 50 else (2000 if total_deals >= 30 else 0)
    
    final_total = d_total + c_total + l_total + u_total + monthly_bonus + b_total
    return final_total, total_deals, monthly_bonus, l_total, d_total, u_total, b_total, b_note

# ========================
# 產生無顏色格式化 Excel 的函式
# ========================
def generate_plain_excel(meta_data, total_v, result, deal_dict, classes, loyalty_dict, upgrades, d_bonus, l_bonus, u_bonus, m_bonus, b_bonus, b_note, emp_type, b_count):
    # 基本資訊列
    header_info = pd.DataFrame([meta_data])
    
    # 左側資料：統計
    stats_data = {
        "統計項目": ["總轉換筆數", "預計總獎金", "體驗成交(筆)", "補開課程(次)", "回流人數(人)", "結構升級(次)", "品牌推廣(位)"],
        "數據": [total_v, result, sum(deal_dict.values()), classes, sum(loyalty_dict.values()), sum(upgrades.values()), b_count]
    }
    
    # 右側資料：明細
    details_data = {
        "獎金明細項目": ["員工身份", "體驗成交獎金", "補位獎金", "回流獎金", "結構升級獎金", "品牌知名度獎金", "月高手獎勵", "獎金說明備註"],
        "金額/內容": [emp_type, d_bonus, classes * 30, l_bonus, u_bonus, b_bonus, m_bonus, b_note]
    }
    
    df_stats = pd.DataFrame(stats_data)
    df_details = pd.DataFrame(details_data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 1. 寫入基本資訊 (Row 1)
        header_info.to_excel(writer, index=False, sheet_name='獎金結算', startrow=0, startcol=0)
        
        # 2. 寫入統計與明細 (Row 4 開始)
        df_stats.to_excel(writer, index=False, sheet_name='獎金結算', startrow=3, startcol=0)
        df_details.to_excel(writer, index=False, sheet_name='獎金結算', startrow=3, startcol=3)
        
        workbook = writer.book
        worksheet = writer.sheets['獎金結算']
        
        # 樣式定義 (僅保留字體與對齊)
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        red_font = Font(color="FF0000", bold=True)

        # 格式化標題列 (無顏色，僅粗體)
        # 第一排基本資訊標題
        for col in range(1, 4):
            worksheet.cell(row=1, column=col).font = bold_font
            worksheet.cell(row=1, column=col).alignment = center_align
            worksheet.cell(row=2, column=col).alignment = center_align

        # 第四排報表統計標題
        for col in [1, 2, 4, 5]:
            worksheet.cell(row=4, column=col).font = bold_font
            worksheet.cell(row=4, column=col).alignment = center_align

        # 調整欄寬與負數標紅
        for col_idx in range(1, 6):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 20
            for row in range(4, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.alignment = center_align
                # 負數數值標紅
                if col_idx in [2, 5] and isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.font = red_font

    return output.getvalue()

# ========================
# Streamlit UI (直式排版)
# ========================
st.set_page_config(page_title="業務獎金計算系統", layout="centered")

st.title("業務獎金計算系統")

# --- 0. 基本資訊區 ---
st.header("基本資訊")
branch = st.selectbox("所屬館別", ["義昌館", "高美館", "中山館", "巨蛋館"])
editor_name = st.text_input("小編姓名", placeholder="請輸入姓名")
report_date = st.date_input("報表日期", datetime.date.today())

# --- 1. 體驗成交與品牌推廣 ---
st.header("1. 體驗與品牌推廣")
d0 = st.number_input("當天成交(筆)", min_value=0, step=1)
d12 = st.number_input("48小時(筆)", min_value=0, step=1)
d37 = st.number_input("7天內(筆)", min_value=0, step=1)

emp_type = st.radio("員工身份", ["正職人員", "兼職人員"], horizontal=True)
brand_count = st.number_input("知名度推廣人數 (位)", min_value=0, step=1)

# --- 2. 課程與回流 ---
st.header("2. 課程與回流獎金")
classes = st.number_input("補開課次數", min_value=0, step=1)

st.subheader("加發回流購課獎金 (人)")
l_c1, l_c2, l_c3, l_c4 = st.columns(4)
with l_c1: l10 = st.number_input("10堂", min_value=0, step=1)
with l_c2: l20 = st.number_input("20堂", min_value=0, step=1)
with l_c3: l30 = st.number_input("30堂", min_value=0, step=1)
with l_c4: l40 = st.number_input("40堂", min_value=0, step=1)

# --- 3. 結構升級 ---
st.header("3. 結構升級獎")
u1 = st.number_input("1對2 變 1對3以上 (次)", min_value=0, step=1)
u2 = st.number_input("團課 變 期班 (次)", min_value=0, step=1)
u3 = st.number_input("包班成立 (次)", min_value=0, step=1)

# --- 計算按鈕 ---
st.divider()
if st.button("開始計算並生成報表", type="primary", use_container_width=True):
    if not editor_name:
        st.error("請輸入小編姓名再進行計算")
    else:
        # 觸發氣球特效
        st.balloons()
        
        deal_dict = {"當天": d0, "48小時": d12, "7天內": d37}
        loyalty_dict = {"10堂": l10, "20堂": l20, "30堂": l30, "40堂": l40}
        upgrades = {"1對2變1對3": u1, "團課變期班": u2, "包班成立": u3}
        
        (result, total_v, m_bonus, l_bonus, d_bonus, u_bonus, b_bonus, b_note) = calculate_bonus(
            deal_dict, classes, loyalty_dict, upgrades, (emp_type == "正職人員"), brand_count
        )
        
        # 整理 Metadata
        meta_data = {
            "館別": branch,
            "小編姓名": editor_name,
            "報表日期": str(report_date)
        }
        
        # 產生無顏色 Excel
        excel_data = generate_plain_excel(
            meta_data, total_v, result, deal_dict, classes, loyalty_dict, 
            upgrades, d_bonus, l_bonus, u_bonus, m_bonus, b_bonus, b_note, emp_type, brand_count
        )
        
        st.success(f"計算完成。總獎金：NT$ {result}")
        
        st.download_button(
            label=f"下載 {report_date} 結算報表",
            data=excel_data,
            file_name=f"{branch}_{editor_name}_{report_date}_獎金結算.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
